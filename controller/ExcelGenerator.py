import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from model.Model import Model
from openpyxl.styles import Font
import pymannkendall as mk

class ExcelGenerator:
    model = Model()
    wb = Workbook()
    ws = wb.active
    missing_consecutives_months = 0
    start_month = None
    start_year = None

    def get_original_raw_data(self, station_name):
        select = f"SELECT ESTCLAVE FROM ESTACIONES WHERE ESTNOMBRE = '{station_name}'"
        self.model.cursor.execute(select)
        estclave = self.model.cursor.fetchone()[0]

        select = f"""
                    SELECT EXTRACT(YEAR FROM fecha) AS year, 
                        EXTRACT(MONTH FROM fecha) AS month, 
                        SUM(precipitacion) AS total_precipitacion, 
                        COUNT(CASE WHEN precipitacion IS NOT NULL THEN 1 END) AS registros
                    FROM registrodiario
                    WHERE ESTCLAVE = {estclave}
                    GROUP BY EXTRACT(YEAR FROM fecha), EXTRACT(MONTH FROM fecha)
                    ORDER BY EXTRACT(YEAR FROM fecha), EXTRACT(MONTH FROM fecha);
        """
        self.model.cursor.execute(select)
        raw_data = self.model.cursor.fetchall()

        self.start_month = raw_data[0][1]
        self.start_year= raw_data[0][0]

        return raw_data, estclave
    
    def get_modified_data(self, station_name, n):
        select = f"SELECT ESTCLAVE FROM ESTACIONES WHERE ESTNOMBRE = '{station_name}'"
        self.model.cursor.execute(select)
        estclave = self.model.cursor.fetchone()[0]

        select = f"select * from get_precipitation_estimation({estclave}, {n});"
        self.model.cursor.execute(select)
        raw_data = self.model.cursor.fetchall()

        return raw_data, estclave
    
    def generate_trend_analysis(self, station, data_type, n=0):
        raw_data = None

        if data_type == 'original_data':
            raw_data = self.get_original_raw_data(station)
        else:
            raw_data = self.get_modified_data(station, n)

        self.ws.title = "Analisis de Tendencia"

        self.ws['B3'] = 'Análisis de tendencia: ' + str(station) + ' (' + str(raw_data[1]) + ')'
        self.ws['B3'].font = Font(name='Arial', size=14, bold=True, color="000000")

        self.ws['B5'] = 'Datos crudos'
        self.ws['B5'].font = Font(name='Arial', size=12, bold=True, color="000000")

        self.ws['H5'] = 'Promedio móvil'
        self.ws['H5'].font = Font(name='Arial', size=12, bold=True, color="000000")

        self.ws['B7'] = 'Año'
        self.ws['C7'] = 'Mes'
        self.ws['D7'] = 'Sum-Precipitacion'
        self.ws['E7'] = 'N'

        data = {
            'month': [],
            'value': []
        }
        
        max_month = 0
        max_sum = 0
        self.missing_consecutives_months = 0    
        missing = 0
        generated_data = 0

        for i, r in enumerate(raw_data[0], start=8):
            self.ws[f'B{i}'] = r[0]
            self.ws[f'C{i}'] = r[1]
            if r[3] > 15 or data_type == 'modified_data':
                self.ws[f'D{i}'] = r[2]
            self.ws[f'E{i}'] = r[3]
                
            if(r[2] == None or r[3] < 15):
                if(data_type != 'original_data'):
                    generated_data+=1

                missing += 1
                if (missing > 6):
                    self.missing_consecutives_months = max(self.missing_consecutives_months, missing)                    
                continue
            missing = 0

            data['month'].append(i-7)            
            data['value'].append(r[2])
            max_month = i-7
            max_sum = max(max_sum, r[2])

        self.generate_scatter_plot(data, raw_data[1], max_month, max_sum)
        
        data, max_month, max_value = self.prepare_rolling_mean_data(12, raw_data[0], 'H')
        self.generate_rolling_mean_plot(data, 12, max_month, max_value, raw_data[1])
        
        data, max_month, max_value = self.prepare_rolling_mean_data(18, raw_data[0], 'I')
        self.generate_rolling_mean_plot(data, 18, max_month, max_value, raw_data[1])

        scatter_plot_img = Image(f'./imgs/scatter_plot_{raw_data[1]}.png')
        rolling_mean_12 = Image(f'./imgs/rolling_mean_12_{raw_data[1]}.png')
        rolling_mean_18 = Image(f'./imgs/rolling_mean_18_{raw_data[1]}.png')

        self.ws.add_image(scatter_plot_img, 'P8')  
        self.ws.add_image(rolling_mean_12, 'P33')
        self.ws.add_image(rolling_mean_18, 'AA33')

        self.generate_mannkendall_test(data['value'])

        self.ws['K5'] = f'Datos generados: {generated_data}'
        self.ws['K5'].font = Font(name='Arial', size=12, bold=True, color="000000")

        if data_type == 'original_data':
            self.wb.save(f'./excels/analisis_tendencia_precipitacion_{station}.xlsx')     
        else:
            self.wb.save(f'./excels/analisis_tendencia_precipitacion_modificado_{station}.xlsx')     
             
    def generate_mannkendall_test(self, data, row = 10):
        self.ws[f'K{row}'] = 'Trend'
        self.ws[f'K{row+1}'] = 'H'
        self.ws[f'K{row+2}'] = 'P'
        self.ws[f'K{row+3}'] = 'Z'
        self.ws[f'K{row+4}'] = 'Tau'
        self.ws[f'K{row+5}'] = 'S'
        self.ws[f'K{row+6}'] = 'Var_S'
        self.ws[f'K{row+7}'] = 'Slope'
        self.ws[f'K{row+8}'] = 'Intercept'

        result = mk.original_test(data)
        self.ws[f'L{row}'] = result.trend
        self.ws[f'L{row+1}'] = result.h
        self.ws[f'L{row+2}'] = result.p
        self.ws[f'L{row+3}'] = result.z
        self.ws[f'L{row+4}'] = result.Tau
        self.ws[f'L{row+5}'] = result.s
        self.ws[f'L{row+6}'] = result.var_s
        self.ws[f'L{row+7}'] = result.slope
        self.ws[f'L{row+8}'] = result.intercept

        result = mk.seasonal_test(data)
        self.ws[f'M{row}'] = result.trend
        self.ws[f'M{row+1}'] = result.h
        self.ws[f'M{row+2}'] = result.p
        self.ws[f'M{row+3}'] = result.z
        self.ws[f'M{row+4}'] = result.Tau
        self.ws[f'M{row+5}'] = result.s
        self.ws[f'M{row+6}'] = result.var_s
        self.ws[f'M{row+7}'] = result.slope
        self.ws[f'M{row+8}'] = result.intercept

        result = mk.sens_slope(data)
        self.ws[f'N{row+7}'] = result.slope
        self.ws[f'N{row+8}'] = result.intercept


    def generate_scatter_plot(self, data, estclave, max_month, max_value):
        df = pd.DataFrame(data)

        plt.scatter(df['month'], df['value'], color='blue', alpha=0.5, label='Sum-Precipitacion')
        df['value'] = df['value'].astype(float)

        trend = np.polyfit(df['month'], df['value'], 1)
        trend_line = np.poly1d(trend)

        plt.plot(df['month'], trend_line(df['month']), color='red', linewidth=1, label='Trend Line')
        plt.text(float(max_month)*0.65, float(max_value)*0.80,  ("y = %.6fx+(%.6f)"%(trend[0],trend[1])), fontsize = 8, color='red')
        
        for i in range(len(df) - 1):
                        plt.plot([df['month'].iloc[i], df['month'].iloc[i + 1]], 
                        [df['value'].iloc[i], df['value'].iloc[i + 1]], 
                        color='blue', linestyle='dashed', linewidth=1)

        plt.xlabel("Meses")
        plt.ylabel("Suma precipitacion")
        plt.title("Serie de tiempo")
        plt.legend()

        plt.savefig(f'./imgs/scatter_plot_{estclave}.png')
        plt.close()


    def prepare_rolling_mean_data(self, window, raw_data, letter):
        data = {'month': [], 'value': []}
        sum = 0.0
        n = 0

        max_month = 0
        max_value = 0
        
        for i in range(window):
            if raw_data[i][2] is not None:
                sum += float(raw_data[i][2])
                n += 1
        
        l = 0  
        for i in range(window, len(raw_data) - window):
            if n <= 0:
                break

            average = sum / n
            max_value = max(average, max_value)
            max_month = i + 1

            data['value'].append(average)
            data['month'].append(i + 1)

            self.ws[f'{letter}{int(i - window + (8 + window / 2 - 1))}'] = average

            if raw_data[l][2] is not None:
                n -= 1
                sum -= float(raw_data[l][2])
            if raw_data[i][2] is not None:
                n += 1
                sum += float(raw_data[i][2])
            l += 1

        return data, max_month, max_value

    def generate_rolling_mean_plot(self, data, window, max_month, max_value, estclave, interval="", start=0):
        df = pd.DataFrame(data)
        df['rolling-window'] = df['value'].astype(float)

        trend = np.polyfit(df['month'], df['rolling-window'], 1)
        trend_line = np.poly1d(trend)

        plt.plot(df['month'], trend_line(df['month']), color='red', linewidth=1, label='Trend Line')
        plt.text(((start/2)+max_month) * 0.65, max_value * 0.8, f"y = {trend[0]:.6f}x + ({trend[1]:.6f})", fontsize=8, color='red')
        plt.plot(df['month'], df['rolling-window'], label=f'Promedio móvil {window} meses', linestyle='solid')

        plt.xlabel('Mes')
        plt.ylabel('Precipitacion')
        plt.title(f'Promedio móvil {window} meses')
        plt.legend()

        plt.savefig(f'./imgs/rolling_mean_{window}_{estclave}{interval}.png')
        plt.close()

    def reset_report(self):
        self.wb = Workbook()
        self.ws = self.wb.active