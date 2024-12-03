from controller.Controller import ExcelGenerator
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import pymannkendall as mk
import pandas as pd

class TrendFinder:
    reporter = ExcelGenerator()
    
    def search_trend(self, interval, station_name):
        self.reporter.reset_report()
        raw_data, estclave = self.reporter.get_original_raw_data(station_name)
        rolling_mean_data, max_month, max_value = self.reporter.prepare_rolling_mean_data(12, raw_data, 'C')

        self.reporter.ws['B3'] = f'Periodos con tendencia de tamaño {interval} meses: ' + str(station_name) + ' (' + str(estclave) + ')'
        self.reporter.ws['B3'].font = Font(name='Arial', size=14, bold=True, color="000000")

        trend_intervals = 0
        last_graph = 10
        data_length = len(rolling_mean_data['month'])
        interval += 1

        for start in range(0, data_length, interval):  
            end = min(start + interval, data_length) 
            data = {
                'month': rolling_mean_data['month'][start:end],
                'value': rolling_mean_data['value'][start:end]
            }
            
            if len(data['value']) < interval:  
                break
            
            df = pd.DataFrame(data)
            df['rolling-window'] = df['value'].astype(float)
            mankendall = mk.original_test(df['rolling-window'])

            if mankendall.trend != 'no trend':
                self.reporter.generate_rolling_mean_plot(data, 12,  data['month'][-1], 40, estclave, f"_interval_{data['month'][0]}", data['month'][0])
                rolling_mean_12 = Image(f'./imgs/rolling_mean_12_{estclave}_interval_{data['month'][0]}.png')
                self.reporter.ws.add_image(rolling_mean_12, f'P{last_graph}')
                self.reporter.generate_mannkendall_test(df['rolling-window'], last_graph)

                real_year, real_month = self.calculate_date(self.reporter.start_year, self.reporter.start_month, data['month'][0])
                self.reporter.ws[f'P{last_graph-2}'] = 'Inicio del periodo: ' 
                self.reporter.ws[f'P{last_graph-1}'] = 'Fin del periodo: ' 

                self.reporter.ws[f'Q{last_graph-2}'] = f'{real_year}-{real_month}'
                real_year, real_month = self.calculate_date(self.reporter.start_year, self.reporter.start_month, data['month'][-1])
                self.reporter.ws[f'Q{last_graph-1}'] = f'{real_year}-{real_month}'

                last_graph += 27
                trend_intervals += 1

        if trend_intervals > 0:
            self.reporter.wb.save(f'./excels/busqueda_tendencia_intervalo{interval-1}_{station_name}.xlsx')

        return trend_intervals
    
    def calculate_date(self, start_year, start_month, months_to_add):
        total_months = start_month + months_to_add
        new_year = start_year + (total_months - 1) // 12
        new_month = (total_months - 1) % 12 + 1
        return new_year, new_month